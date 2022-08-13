from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

import inspect
import sys
import time
import os.path
from .models import Job, JobFurnance, JobSetting
from master.models import PartMF, Customer, ConstantMF, Furnance
from openpyxl.drawing.image import Image

SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
from tohken.settings import BASE_DIR
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import Alignment
import qrcode
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

def whoami():
    frame = inspect.currentframe()
    return inspect.getframeinfo(frame).function


def cell_val(o, letter):
    index = column_index_from_string(letter)
    v = o[index-1].value
    return v

templateMap = {
        "conveyor": "Conveyor",
        "batch-tvq": "Batch TVQ",
        "batch-tvt": "Batch TVT",
        "batch-tfq_tft": "Batch TFQ, TFT",
        "batch-tgc_trgc": "Batch TGC, TRGC",
        "batch-tbt": "Batch TBT",
        "batch-tih": "Batch TIH",
        "batch-tsz": "Batch TSZ",
        "batch-tsp_tsap": "Batch TSP-TSAP",
        "set-conveyor": "Set conveyor",
        "set-batch": "Set Batch",
        "tsh-drum": "TSH Drum",
        "tsh-hanger": "TSH Hanger",
        "inspection": "INSPECTION"
        }

def generateExportFile(pk):
    job  = Job.objects.get(pk=pk)
    print('glass' in job.safety)
    wb = load_workbook('./testtemplate2.xlsx')

    ut = int(time.time())
    codeno = job.part.partcode
    #ws = wb['Standard']
    #ws['E6'] = job.part.partcode
    for k,v in templateMap.items():
        print(v)
        print(wb.sheetnames)
        if v in wb.sheetnames:
            print("hidden mode")
            wb[v].sheet_state = "hidden"

    list_jf = []
    for jf in job.jobfurnance_set.all():
            if jf.enable == True and jf.furnance:
                list_jf.append(str(jf.furnance))
                generateStd(wb, job,  jf)

    generateInspection(wb, job)
    generateSetting(wb, job)
    generateShotBlast(wb, job)
    filename = "{0}_{2}_{1}.xlsx".format(codeno.replace("-", ""),"_".join(list_jf),  ut)
    print(filename)
    outfile = '/media/exports/'+filename
    wb.save(filename=BASE_DIR+outfile)
    return outfile

def generateSetting(wb,job):
    jobSettings = job.jobsetting_set.all()


    for js in jobSettings:
        if js.setting_type == "Batch":
            generateSetBatch(wb, js)
        elif js.setting_type == "Conveyor":
            generateSetConveyor(wb, js)

def generateShotBlast(wb,job):
    jobSettings = job.shotblast_set.all()

    for js in jobSettings:
        if js.sb_type == "Drum":
            generateSBDrum(wb, js)
        elif js.sb_type == "Hanger":
            generateSBHanger(wb, js)


def sheetInfo(ws, job, values):

    ws['E5'] = job.part.partcode
    ws["ac3"] =  job.version_code
    ws["ac2"] = job.updated_at

    part = job.part
    photos = part.partphoto_set.all()
    if len(photos) > 0:
        img = Image(f"{BASE_DIR}{photos[0].file.url}")
        ws.add_image(img, "Y12")

    customer = Customer.objects.get(code = part.custom_no)
    ws['Q4'] = "*{0}*".format(part.partcode)
    ws['E4'] = part.doc_code
    ws['E3'] = job.part_rank
    ws["E6"] = customer.name
    ws["E7"] = part.designno
    ws["E8"] = part.partname
    ws["E9"] = part.size
    ws["E10"] = part.treatcd
    ws["E11"] = part.mat
    ws["Q11"] = part.kgperpc
    ws["B13"] = part.shtester1
    ws["E13"] = part.low1
    ws["I13"] = part.up1

    furnance = values.get('furnance_jig', '')

    doc_code = part.doc_code
    if doc_code is None:
        doc_code = "000"

    qrData = f"{part.doc_code}|{part.partname}|{part.designno}|{doc_code[0:3].zfill(4)}:{customer.name}|{furnance}|{job.updated_at.strftime('%d-%B-%y')}|{str(job.version_code).zfill(2)}"
    print(f"qr = {qrData}")

    img = qrcode.make(qrData)
    thimg = img.resize((100, 100))
    ts = time.time()
    filename =  f"{int(ts)}.png"
    print(filename)
    fullPath = BASE_DIR + "/media/" + filename
    thimg.save(fullPath, format='PNG')
    qrimg = Image(fullPath)
    ws.add_image(qrimg, "Y5")
    '''
    if part.shtester1 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.shtester1. class_code=7)
            ws["B14"] = const.data1
        except:
            pass
    '''

    if part.shtester2 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.shtester2, class_code=7)
            ws["B14"] = const.data1
            ws["E14"] = part.low2
            ws["I14"] = part.up2
        except:
            pass

    if part.shtester3 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.shtester3, class_code=7)
            ws["B15"] = const.data1
            ws["E15"] = part.low3
            ws["I15"] = part.up3
        except ConstantMF.DoesNotExist:
            pass

    if part.chtester1 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.chtester1, class_code = 7)
            ws["N13"] = const.data1
            ws["Q13"] = part.chlow1
            ws["U13"] = part.chup1
        except ConstantMF.DoesNotExist:
            pass

    if part.chtester2 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.chtester2, class_code = 7)
            ws["N14"] = const.data1
            ws["Q14"] = part.chlow2
            ws["U14"] = part.chup2
        except ConstantMF.DoesNotExist:
            pass

    if part.othertester is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.othertester, class_code = 7)
            ws["Q15"] = "{0} {1}-{2}".format(const.data1, part.olow1, part.oup1)
        except ConstantMF.DoesNotExist:
            pass

    if part.effpoint1 is not None and part.effpoint1 != 0:
        try:
            effpoint  = ConstantMF.objects.get(data_code = part.effpoint1, class_code = 8)
            ws["E17"] = effpoint.data
            ws["I17"] = part.effcriterion1
            ws["M17"] = part.eff_low1
            ws["Q17"] = part.eff_up1
        except ConstantMF.DoesNotExist:
            pass

    if part.effpoint2 is not None and part.effpoint2 != 0:
        try:
            effpoint2  = ConstantMF.objects.get(data_code = part.effpoint2, class_code = 8)
            ws["E18"] = effpoint2.data
            ws["I18"] = part.effcriterion2
            ws["M18"] = part.eff_low2
            ws["Q18"] = part.eff_up2
        except ConstantMF.DoesNotExist:
            pass

    ws["E19"] = part.total_measures
    ws["I19"] = part.total_low


    if values is not None:
        if 'safety' in values:
            ws['E20'] = 'R' if 'helmet' in values['safety'] else ''
            ws['I20'] = 'R' if 'glass' in values['safety'] else ''
            ws['M20'] = 'R' if 'ear_fill' in values['safety'] else ''
            ws['Q20'] = 'R' if 'nose' in values['safety'] else ''
            ws['U20'] = 'R' if 'eam' in values['safety'] else ''
            ws['X20'] = 'R' if 'arm' in values['safety'] else ''
            ws['AB20'] = 'R' if 'back' in values['safety'] else ''

            ws['E21'] = 'R' if 'safety_shoes' in values['safety'] else ''
            ws['I21'] = 'R' if 'ear_cover' in values['safety'] else ''
            ws['M21'] = 'R' if 'belt' in values['safety'] else ''
            ws['Q21'] = 'R' if 'glove' in values['safety'] else ''
            ws['U21'] = 'R' if 'glove_c' in values['safety'] else ''
            ws['X21'] = 'R' if 'other' in values['safety'] else ''

def generateSetConveyor(wb, jobSetting):
    ws = wb["Set conveyor"]
    ws.sheet_state = "visible"
    keyMap = {
            'furnance': 'f48',
            'coveyor_length': 'q48',
            'speed': 'z48',
            'part_kgperpc': 'f49',
            'part_amt': 'q49',
            'total_part_weight': 'z49',
            'ami_wpc': 'f50',
            'ami_amt': 'q50',
            'total_ami_weight': 'z50',
            'nrow_hr': 'f51',
            'length_row': 'q51',
            'total_row_hr': 'z51',
            'qty_hr': 'z52',
            'limit_weight':  'f53',
            'total_weight': 'z53',
            'setting_method': 'b42',
            'comment': 'b56',

            }
    print(jobSetting.values)

    job = jobSetting.job

    if job.status == "draft":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET ( Draft )"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "trial":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "trial_approve":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "standard":
        ws["E1"] = "SETTING JOB STANDARD"
        ws["A1"] = "WI-PD-003"
        x53 = ws["X61"]
        x53.value = ""
        x53.border = None
        #aa53 = ws["AA53"]
        #aa53.value = ""
        #aa53.border = None

    thin = Side(border_style="thin", color="000000")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin,
            diagonal=Side(border_style="thin",color='000000'),
            diagonal_direction=1,
            diagonalDown=True,
            )



    for c in ["X66", "AA66", "AD66"]:
        if job.status == "draft":
            ws[c].border  = border
        if job.status == "trial":
            ws[c].border  = border
        if job.status == "trial_approve":
            print("diagonalUp = True")
            ws[c].border  = border
        if job.status == "standard":
            ws[c].border  = None

    ws['b42'].alignment = Alignment(wrapText=True)
    ws['b56'].alignment = Alignment(wrapText=True)
    sheetInfo(ws, jobSetting.job, jobSetting.values)
    for k,v in keyMap.items():
        ws[v] = jobSetting.values.get(k, 0.0)

    settingPhotos = list(jobSetting.settingphoto_set.filter(group = "Setting Method"))

    photoCell = ["a23", "i23", "q23", "y23", "a31", "i31", "q31", "y31"]
    captionCell = ["a30", "i30", "q30", "y30", "a38", "i38", "q38", "y38"]
    if settingPhotos:
        for i, v in enumerate(settingPhotos):
            bigPic = v
            if bigPic and bigPic.file:
                img = Image(f"{BASE_DIR}{bigPic.file.url}")
                ws.add_image(img, photoCell[i])
                ws[captionCell[i]] = bigPic.caption

    for i in range(66, 67):
        print("range ...")
        print(f"A{i} = {job.version_code } C{i} = {job.updated_at}")
        ws[f"A{i}"] = job.version_code
        ws[f"C{i}"] = job.updated_at
        ws[f"F{i}"] = job.rev_note
        ws[f"U{i}"] = job.c_no

    if 'parent' in job.meta_data:
        if len(job.meta_data['parent']) >= 1:
            p0 = job.meta_data['parent'][0]
            print(f"parint {job.meta_data['parent'][0]}")
            #print(f"parint {job.parent}")
            if p0:
                temp = Job.objects.get(pk = p0['id'])
                ws["A67"] = temp.version_code
                ws["C67"] = temp.updated_at
                ws["F67"] = temp.rev_note
                ws["U67"] = temp.c_no

        if len(job.meta_data['parent']) >= 2:
            p1 = job.meta_data['parent'][1]
            print(f"parint {job.meta_data['parent'][0]}")
            #print(f"parint {job.parent}")
            if p1:
                temp = Job.objects.get(pk = p1['id'])
                ws["A68"] = temp.version_code
                ws["C68"] = temp.updated_at
                ws["F68"] = temp.rev_note
                ws["U68"] = temp.c_no

def generateSBDrum(wb, jobSetting):
    ws = wb["TSH Drum"]
    ws.sheet_state = "visible"
    keyMap = {
            'furnance': 'e23',
            'mediaSize': 'r23',
            'ptn_no': 'e24',
            'time_op': 'n24',
            'kg_per_time': 'z24',
            }

    sheetInfo(ws, jobSetting.job, jobSetting.values)
    for k,v in keyMap.items():
        ws[v] = jobSetting.values.get(k, 0.0)

    ws['l51'] = jobSetting.values.get('ptn_no', '-')


    jigPhotos = list(jobSetting.sbphoto_set.filter(group = "Setting"))
    settingPhotos = list(jobSetting.sbphoto_set.filter(group = "Usage"))

    jigPhotosCellMap = (("b27", "b34") , ("h27", "h34"), ("n27", "n34"), ("t27", "t34"), ("z27", "z34"))


    allCells = ("b27", "h27", "n27", "t27", "z27", "a40", "l40", "q40", "y40")
    for index, val in enumerate(allCells):
            img = Image(f"{BASE_DIR}/static/imgs/tshdefault/drum{index+1}.png")
            img.width = 168
            img.height = 110
            ws.add_image(img, allCells[index].upper())

    for index, val in enumerate(jigPhotos):
        if val.file:
            print(f"val = {val}, {jigPhotosCellMap[index][0]}, {jigPhotosCellMap[index][1]} {val.file.url} {val.caption}")

            img = Image(f"{BASE_DIR}{val.file.url}")
            ws.add_image(img, jigPhotosCellMap[index][0].upper())
            ws[jigPhotosCellMap[index][1]] = val.caption


    settingPhotosCellMap = ("a40", "l40", "q40", "y40")


    for index, val in enumerate(settingPhotos):
        if val.file:
            img = Image(f"{BASE_DIR}{val.file.url}")
            ws.add_image(img, settingPhotosCellMap[index].upper())

    job = jobSetting.job

    if job.status == "draft":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET ( Draft )"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "trial":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "trial_approve":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "standard":
        ws["E1"] = "SETTING JOB STANDARD"
        ws["A1"] = "WI-PD-003"
        x53 = ws["X61"]
        x53.value = ""
        x53.border = None
        #aa53 = ws["AA53"]
        #aa53.value = ""
        #aa53.border = None

    thin = Side(border_style="thin", color="000000")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin,
            diagonal=Side(border_style="thin",color='000000'),
            diagonal_direction=1,
            diagonalDown=True,
            )



    for c in ["X66", "AA66", "AD66"]:
        if job.status == "draft":
            ws[c].border  = border
        if job.status == "trial":
            ws[c].border  = border
        if job.status == "trial_approve":
            print("diagonalUp = True")
            ws[c].border  = border
        if job.status == "standard":
            ws[c].border  = None

    for i in range(66, 67):
        print("range ...")
        print(f"A{i} = {job.version_code } C{i} = {job.updated_at}")
        ws[f"A{i}"] = job.version_code
        ws[f"C{i}"] = job.updated_at
        ws[f"F{i}"] = job.rev_note
        ws[f"U{i}"] = job.c_no

    if len(job.meta_data['parent']) >= 1:
        p0 = job.meta_data['parent'][0]
        print(f"parint {job.meta_data['parent'][0]}")
        #print(f"parint {job.parent}")
        if p0:
            temp = Job.objects.get(pk = p0['id'])
            ws["A67"] = temp.version_code
            ws["C67"] = temp.updated_at
            ws["F67"] = temp.rev_note
            ws["U67"] = temp.c_no

    if len(job.meta_data['parent']) >= 2:
        p1 = job.meta_data['parent'][1]
        print(f"parint {job.meta_data['parent'][0]}")
        #print(f"parint {job.parent}")
        if p1:
            temp = Job.objects.get(pk = p1['id'])
            ws["A68"] = temp.version_code
            ws["C68"] = temp.updated_at
            ws["F68"] = temp.rev_note
            ws["U68"] = temp.c_no

def generateSBHanger(wb, jobSetting):
    ws = wb["TSH Hanger"]
    ws.sheet_state = "visible"
    keyMap = {
            'furnance': 'e23',
            'mediaSize': 'r23',
            'ptn_no': 'e24',
            'time_op': 'n24',
            'kg_per_time': 'z24',
            'no_per_level': 'p35',
            'pc_per_br': 'p36',
            'use_level': 'p37',
            'total_pcs': 'p38',
            'time_set': 'l51'
            }

    sheetInfo(ws, jobSetting.job, jobSetting.values)
    for k,v in keyMap.items():
        print(f"{k} = {v}")
        ws[v] = jobSetting.values.get(k, 0.0)



    jigPhotos = list(jobSetting.sbphoto_set.filter(group = "Setting"))
    settingPhotos = list(jobSetting.sbphoto_set.filter(group = "Usage"))
    print(jigPhotos)
    print(settingPhotos)
    jigPhotosCellMap = ("b27" , "h27", "n27", "t27", "z27")

    allCells = ("b27", "h27", "n27", "t27", "z27", "a40", "l40", "q40", "y40")
    for index, val in enumerate(allCells):
            img = Image(f"{BASE_DIR}/static/imgs/tshdefault/hanger{index+1}.png")
            img.width = 168
            img.height = 110
            ws.add_image(img, allCells[index].upper())

    for index, val in enumerate(jigPhotos):
        if val.file:
            print(f"val = {val}, {jigPhotosCellMap[index][0]}, {jigPhotosCellMap[index][1]} {val.file.url} {val.caption}")

            img = Image(f"{BASE_DIR}{val.file.url}")
            ws.add_image(img, jigPhotosCellMap[index].upper())
            #ws[jigPhotosCellMap[index][1]] = val.caption


    settingPhotosCellMap = ("a40", "l40", "q40", "y40")

    for index, val in enumerate(settingPhotos):
        if val.file:
            img = Image(f"{BASE_DIR}{val.file.url}")
            ws.add_image(img, settingPhotosCellMap[index].upper())

    job = jobSetting.job

    if job.status == "draft":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET ( Draft )"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "trial":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "trial_approve":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "standard":
        ws["E1"] = "SETTING JOB STANDARD"
        ws["A1"] = "WI-PD-003"
        x53 = ws["X61"]
        x53.value = ""
        x53.border = None
        #aa53 = ws["AA53"]
        #aa53.value = ""
        #aa53.border = None

    thin = Side(border_style="thin", color="000000")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin,
            diagonal=Side(border_style="thin",color='000000'),
            diagonal_direction=1,
            diagonalDown=True,
            )



    for c in ["X66", "AA66", "AD66"]:
        if job.status == "draft":
            ws[c].border  = border
        if job.status == "trial":
            ws[c].border  = border
        if job.status == "trial_approve":
            print("diagonalUp = True")
            ws[c].border  = border
        if job.status == "standard":
            ws[c].border  = None

    for i in range(66, 67):
        print("range ...")
        print(f"A{i} = {job.version_code } C{i} = {job.updated_at}")
        ws[f"A{i}"] = job.version_code
        ws[f"C{i}"] = job.updated_at
        ws[f"F{i}"] = job.rev_note
        ws[f"U{i}"] = job.c_no

    if len(job.meta_data['parent']) >= 1:
        p0 = job.meta_data['parent'][0]
        print(f"parint {job.meta_data['parent'][0]}")
        #print(f"parint {job.parent}")
        if p0:
            temp = Job.objects.get(pk = p0['id'])
            ws["A67"] = temp.version_code
            ws["C67"] = temp.updated_at
            ws["F67"] = temp.rev_note
            ws["U67"] = temp.c_no

    if len(job.meta_data['parent']) >= 2:
        p1 = job.meta_data['parent'][1]
        print(f"parint {job.meta_data['parent'][0]}")
        #print(f"parint {job.parent}")
        if p1:
            temp = Job.objects.get(pk = p1['id'])
            ws["A68"] = temp.version_code
            ws["C68"] = temp.updated_at
            ws["F68"] = temp.rev_note
            ws["U68"] = temp.c_no

def generateSetBatch(wb, jobSetting):
    ws = wb["Set Batch"]
    ws.sheet_state = "visible"

    job = jobSetting.job

    if job.status == "draft":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET ( Draft )"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "trial":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "trial_approve":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
        ws["X61"] = "Mass Prodution Judgement By"
    if job.status == "standard":
        ws["E1"] = "SETTING JOB STANDARD"
        ws["A1"] = "WI-PD-003"
        x53 = ws["X61"]
        x53.value = ""
        x53.border = None
        #aa53 = ws["AA53"]
        #aa53.value = ""
        #aa53.border = None

    thin = Side(border_style="thin", color="000000")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin,
            diagonal=Side(border_style="thin",color='000000'),
            diagonal_direction=1,
            diagonalDown=True,
            )



    for c in ["X66", "AA66", "AD66"]:
        if job.status == "draft":
            ws[c].border  = border
        if job.status == "trial":
            ws[c].border  = border
        if job.status == "trial_approve":
            print("diagonalUp = True")
            ws[c].border  = border
        if job.status == "standard":
            ws[c].border  = None

    keyMap = {
            'sm_weight': 'g37',
            'sm_amt': 'k37',
            'tray_weight': 'g38',
            'tray_amt': 'k38',
            'wax_weight': 'g39',
            'wax_amt': 'k39',
            'support_ami_weight': 'g40',
            'supporrt_ami_amt': 'k40',
            'plate_weight': 'g41',
            'plate_amt': 'k41',
            'under_ami_weight': 'g42',
            'under_ami_amt': 'k42',
            'cover_ami_weight': 'g43',
            'cover_ami_amt': 'k43',
            'other1_weight': 'g44',
            'other1_amt': 'k44',
            'other2_weight': 'g45',
            'other2_amt': 'k45',
            'other3_weight': 'g46',
            'other3_amt': 'k46',
            'furnance_jig': 'c47',
            'batch_total': 'n47',
            'setting_method': 'b50',
            }
    textMap = {
            'sm_text':  'a37',
            'tray_text': 'a38',
            'wax_text': 'a39',
            'support_ami_text': 'a40',
            'plate_text':'a41',
            'under_ami_text':'a42',
            'cover_ami_text':'a43',
            'other1_text':'a44',
            'other2_text':'a45',
            'other3_text':'a46',
            }
    print(jobSetting.values)
    ws['b50'].alignment = Alignment(wrapText=True)
    sheetInfo(ws, jobSetting.job, jobSetting.values)
    for k,v in keyMap.items():
        ws[v] = jobSetting.values.get(k, 0.0)

    for k,v in textMap.items():
        ws[v] = jobSetting.values.get(k, 0.0)


    jigPhotos = list(jobSetting.settingphoto_set.filter(group = "Jig Setting"))
    settingPhotos = list(jobSetting.settingphoto_set.filter(group = "Setting Method"))

    jigPhotosCellMap = (("a23", "a29") , ("e23", "e29"), ("l23", "l29"), ("m23", "m29"), ("a30", "a36"), ("e30", "e36"), ("l30", "l36"), ("m30", "m36"))

    for index, val in enumerate(jigPhotos):
        if val.file:
            print(f"val = {val}, {jigPhotosCellMap[index][0]}, {jigPhotosCellMap[index][1]} {val.file.url} {val.caption}")

            print(f"{BASE_DIR}{val.file.url}")
            img = Image(f"{BASE_DIR}{val.file.url}")
            ws.add_image(img, jigPhotosCellMap[index][0].upper())
            ws[jigPhotosCellMap[index][1]] = str(val.caption)


    settingPhotosCellMap = ("q23", "y23", "q31", "y31", "q39", "y39")

    for index, val in enumerate(settingPhotos):
        if val.file:
            img = Image(f"{BASE_DIR}{val.file.url}")
            ws.add_image(img, settingPhotosCellMap[index].upper())




    for i in range(66, 67):
        print("range ...")
        print(f"A{i} = {job.version_code } C{i} = {job.updated_at}")
        ws[f"A{i}"] = job.version_code
        ws[f"C{i}"] = job.updated_at
        ws[f"F{i}"] = job.rev_note
        ws[f"U{i}"] = job.c_no

    if len(job.meta_data['parent']) == 1:
        p0 = job.meta_data['parent'][0]
        print(f"parint {job.meta_data['parent'][0]}")
        #print(f"parint {job.parent}")
        if p0:
            temp = Job.objects.get(pk = p0['id'])
            ws["A67"] = temp.version_code
            ws["C67"] = temp.updated_at
            ws["F67"] = temp.rev_note
            ws["U67"] = temp.c_no

    if len(job.meta_data['parent']) == 2:
        p1 = job.meta_data['parent'][1]
        print(f"parint {job.meta_data['parent'][0]}")
        #print(f"parint {job.parent}")
        if p1:
            temp = Job.objects.get(pk = p1['id'])
            ws["A68"] = temp.version_code
            ws["C68"] = temp.updated_at
            ws["F68"] = temp.rev_note
            ws["U68"] = temp.c_no

def generateInspection(wb, job):
    #query job inspection
    inspections = job.jobinspection_set.all()
    if not inspections:
        return

    ws = wb['INSPECTION']

    if job.status == "draft":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET ( Draft )"
    if job.status == "trial":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
    if job.status == "trial_approve":
        ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
        ws["A1"] = "FR-AT-002"
    if job.status == "standard":
        ws["E1"] = "INSPECTION STANDARD"
        ws["A1"] = "WI-QA-067"

    ws.sheet_state = "visible"
    #open inspsection sheet
    #plot data
    startRow = 27
    ws['E5'] = job.part.partcode
    ws["ac3"] =  job.version_code
    ws["ac2"] = job.updated_at

    for i in range(69, 70):
        print("range ...")
        print(f"A{i} = {job.version_code } C{i} = {job.updated_at}")
        ws[f"A{i}"] = job.version_code
        ws[f"C{i}"] = job.updated_at
        ws[f"F{i}"] = job.rev_note
        ws[f"R{i}"] = job.c_no

    job.meta_data['parent'].reverse()
    if len(job.meta_data['parent']) > 0:
        p0 = job.meta_data['parent'][0]
        print(f"parint {job.meta_data['parent'][0]}")
        #print(f"parint {job.parent}")
        if p0:
            temp = Job.objects.get(pk = p0['id'])
            ws["A70"] = temp.version_code
            ws["C70"] = temp.updated_at
            ws["F70"] = temp.rev_note
            ws["R70"] = temp.c_no

        if len(job.meta_data['parent']) >= 2:
            p1 = job.meta_data['parent'][1]
            print(f"parint {job.meta_data['parent'][0]}")
            #print(f"parint {job.parent}")
            if p1:
                temp = Job.objects.get(pk = p1['id'])
                ws["A71"] = temp.version_code
                ws["C71"] = temp.updated_at
                ws["F71"] = temp.rev_note
                ws["R71"] = temp.c_no

    img = qrcode.make('Some data here')
    img.save(BASE_DIR + "/media/" + "opera_house.png", format='PNG')

    part = job.part

    photos = part.partphoto_set.all()
    if len(photos) > 0:
        img = Image(f"{BASE_DIR}{photos[0].file.url}")
        ws.add_image(img, "Y12")

    customer = Customer.objects.get(code = part.custom_no)


    doc_code = part.doc_code
    if doc_code is None:
        doc_code = "000"

    qrData = f"{part.doc_code}|{part.partname}|{part.designno}|{doc_code[0:3].zfill(4)}:{customer.name}|{job.updated_at.strftime('%d-%B-%y')}|{str(job.version_code).zfill(2)}"
    print(f"qr = {qrData}")

    jsty = job.safety_inspection

    ws['E20'] = 'R' if 'helmet' in jsty else ''
    ws['I20'] = 'R' if 'glass' in jsty else ''
    ws['M20'] = 'R' if 'ear_fill' in jsty else ''
    ws['Q20'] = 'R' if 'nose' in jsty else ''
    ws['U20'] = 'R' if 'eam' in jsty else ''
    ws['X20'] = 'R' if 'arm' in jsty else ''
    ws['AB20'] = 'R' if 'back' in jsty else ''

    ws['E21'] = 'R' if 'safety_shoes' in jsty else ''
    ws['I21'] = 'R' if 'ear_cover' in jsty else ''
    ws['M21'] = 'R' if 'belt' in jsty else ''
    ws['Q21'] = 'R' if 'glove' in jsty else ''
    ws['U21'] = 'R' if 'glove_c' in jsty else ''
    ws['X21'] = 'R' if 'other' in jsty else ''

    img = qrcode.make(qrData)
    thimg = img.resize((100, 100))
    ts = time.time()
    filename =  f"{int(ts)}.png"
    print(filename)
    fullPath = BASE_DIR + "/media/" + filename
    thimg.save(fullPath, format='PNG')
    qrimg = Image(fullPath)
    ws.add_image(qrimg, "Y5")

    ws['Q4'] = "*{0}*".format(part.partcode)
    ws['E4'] = part.doc_code
    ws['E3'] = job.part_rank
    ws["E6"] = customer.name
    ws["E7"] = part.designno
    ws["E8"] = part.partname
    ws["E9"] = part.size
    ws["E10"] = part.treatcd
    ws["E11"] = part.mat
    ws["Q11"] = part.kgperpc
    ws["B13"] = part.shtester1
    ws["E13"] = part.low1
    ws["I13"] = part.up1

    #part.doc_code + "|" + part.partname + "|" + part.designno + "|" + job.version_code.zfill(2) + "|"+   "|" + customer.name + "|" + part.doc_code[0:3].zill(4) +

    '''
    if part.shtester1 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.shtester1. class_code=7)
            ws["B14"] = const.data1
        except:
            pass
    '''

    if part.shtester2 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.shtester2, class_code=7)
            ws["B14"] = const.data1
            ws["E14"] = part.low2
            ws["I14"] = part.up2
        except:
            pass

    if part.shtester3 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.shtester3, class_code=7)
            ws["B15"] = const.data1
            ws["E15"] = part.low3
            ws["I15"] = part.up3
        except ConstantMF.DoesNotExist:
            pass

    if part.chtester1 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.chtester1, class_code = 7)
            ws["N13"] = const.data1
            ws["Q13"] = part.chlow1
            ws["U13"] = part.chup1
        except ConstantMF.DoesNotExist:
            pass

    if part.chtester2 is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.chtester2, class_code = 7)
            ws["N14"] = const.data1
            ws["Q14"] = part.chlow2
            ws["U14"] = part.chup2
        except ConstantMF.DoesNotExist:
            pass

    if part.othertester is not None:
        try:
            const = ConstantMF.objects.get(data_code = part.othertester, class_code = 7)
            ws["Q15"] = "{0} {1}-{2}".format(const.data1, part.olow1, part.oup1)
        except ConstantMF.DoesNotExist:
            pass

    if part.effpoint1 is not None and part.effpoint1 != 0:
        try:
            effpoint  = ConstantMF.objects.get(data_code = part.effpoint1, class_code = 8)
            ws["E17"] = effpoint.data
            ws["I17"] = part.effcriterion1
            ws["M17"] = part.eff_low1
            ws["Q17"] = part.eff_up1
        except ConstantMF.DoesNotExist:
            pass

    if part.effpoint2 is not None and part.effpoint2 != 0:
        try:
            effpoint2  = ConstantMF.objects.get(data_code = part.effpoint2, class_code = 8)
            ws["E18"] = effpoint2.data
            ws["I18"] = part.effcriterion2
            ws["M18"] = part.eff_low2
            ws["Q18"] = part.eff_up2
        except ConstantMF.DoesNotExist:
            pass

    ws["E19"] = part.total_measures
    ws["I19"] = part.total_low

    '''
    values = jobfurnance.values
    ws['E20'] = 'R' if 'helmet' in values['safety'] else ''
    ws['I20'] = 'R' if 'glass' in values['safety'] else ''
    ws['M20'] = 'R' if 'ear_fill' in values['safety'] else ''
    ws['Q20'] = 'R' if 'nose' in values['safety'] else ''
    ws['U20'] = 'R' if 'eam' in values['safety'] else ''
    ws['X20'] = 'R' if 'arm' in values['safety'] else ''
    ws['AB20'] = 'R' if 'back' in values['safety'] else ''

    ws['E21'] = 'R' if 'safety_shoes' in values['safety'] else ''
    ws['I21'] = 'R' if 'ear_cover' in values['safety'] else ''
    ws['M21'] = 'R' if 'belt' in values['safety'] else ''
    ws['Q21'] = 'R' if 'glove' in values['safety'] else ''
    ws['U21'] = 'R' if 'glove_c' in values['safety'] else ''
    ws['X21'] = 'R' if 'other' in values['safety'] else ''
    '''

    for f in inspections:
        ws.cell(row = startRow, column = column_index_from_string('B'), value=f.get_inspection_item_display())
        ws.cell(row = startRow, column = column_index_from_string('G'), value=f.get_tester_display() )
        ws.cell(row = startRow, column = column_index_from_string('J'), value=f.sample_size)
        ws.cell(row = startRow + 1, column = column_index_from_string('J'), value=f.sample_points)
        ws.cell(row = startRow + 2, column = column_index_from_string('B'), value=f.sample_note)
        ws.cell(row = startRow , column = column_index_from_string('O'), value=f.standard_note)
        startRow += 3
        #print(f.get_inspection_item_display())

    try:
        img = Image(f"{BASE_DIR}{inspections[0].photo.url}")
        ws.add_image(img, "U23")
    except:
        pass

    try:
        img = Image(f"{BASE_DIR}{inspections[1].photo.url}")
        ws.add_image(img, "U36")
    except:
        pass

    try:
        img = Image(f"{BASE_DIR}{inspections[2].photo.url}")
        ws.add_image(img, "U51")
    except:
        pass

    thin = Side(border_style="thin", color="000000")
    thin = Side(border_style="thin", color="000000")

    border = Border(top=thin, left=thin, right=thin, bottom=thin,
            diagonal=Side(border_style="thin",color='000000'),
            diagonal_direction=1,
            diagonalDown=True,
            )
    #ws.merge_cells('AA69:AC71')
    #waa69 = ws["AA69"]
    #waa69.border = border
    #ws["AA69"].border = border
    #ws["AA69"].value = "OOOOOKKKK"
    #return

    u68 = ws["U68"]
    if job.status == "draft":
        u68.value = "Mass Pro Judgement"
    if job.status == "trial":
        u68.value = "Mass Pro Judgement"
    if job.status == "trial_approve":
        u68.value = "Mass Pro Judgement"
    if job.status == "standard":
        u68.value = "Customer approve"

    for c in ["X69", "AA69", "AD69"]:
        if job.status == "draft":
            ws[c].border  = border
        if job.status == "trial":
            ws[c].border  = border
        if job.status == "trial_approve":
            print("diagonalUp = True")
            ws[c].border  = border
        if job.status == "standard":
            ws[c].border  = None


def generateStd(wb, job, jobfurnance):

    #wb['Conveyor'].sheet_state = 'hidden'
    #wb['Batch TVQ-TFQ'].sheet_state = 'hidden'
    print(f"jobfurnance = {jobfurnance.enable}")

    if jobfurnance.enable  == True:
        templateName = templateMap[jobfurnance.template_name]
        print("templatName "+templateName)
        wb[templateName].sheet_state = "visible"
        ws = wb[templateName]

        if job.status == "draft":
            ws["E1"] = "Draft"
            ws["X53"] = "Mass Prodution Judgement By"
        if job.status == "trial":
            ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
            ws["A1"] = "FR-AT-002"
            ws["X53"] = "Mass Prodution Judgement By"
        if job.status == "trial_approve":
            ws["E1"] = "HEAT TREATMENT OPERATION TRIAL SHEET"
            ws["A1"] = "FR-AT-002"
            ws["X53"] = "Mass Prodution Judgement By"
        if job.status == "standard":
            ws["E1"] = "HEAT TREATMENT JOB STANDARD"
            ws["A1"] = "WI-PD-001"
            x53 = ws["X53"]
            x53.value = ""
            x53.border = None
            aa53 = ws["AA53"]
            aa53.value = ""
            aa53.border = None


        thin = Side(border_style="thin", color="000000")
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin,
                diagonal=Side(border_style="thin",color='000000'),
                diagonal_direction=1,
                diagonalDown=True,
                )



        for c in ["X69", "AA69", "AD69"]:
            if job.status == "draft":
                ws[c].border  = border
            if job.status == "trial":
                ws[c].border  = border
            if job.status == "trial_approve":
                print("diagonalUp = True")
                ws[c].border  = border
            if job.status == "standard":
                ws[c].border  = None

        ws['E5'] = job.part.partcode
        ws["ac3"] =  job.version_code

        part = job.part

        photos = part.partphoto_set.all()
        if len(photos) > 0:
            img = Image(f"{BASE_DIR}{photos[0].file.url}")
            ws.add_image(img, "Y12")

        customer = Customer.objects.get(code = part.custom_no)
        ws['Q4'] = "*{0}*".format(part.partcode)
        ws['E4'] = part.doc_code
        ws['E3'] = job.part_rank
        ws["E6"] = customer.name
        ws["E7"] = part.designno
        ws["E8"] = part.partname
        ws["E9"] = part.size
        ws["AC2"] = job.updated_at
        ws["E3"] = job.part_rank
        ws["E10"] = part.treatcd
        ws["E11"] = part.mat
        ws["Q11"] = part.kgperpc
        ws["B13"] = part.shtester1
        ws["E13"] = part.low1
        ws["I13"] = part.up1

        for i in range(69, 70):
            print("range ...")
            print(f"A{i} = {job.version_code } C{i} = {job.updated_at}")
            ws[f"A{i}"] = job.version_code
            ws[f"C{i}"] = job.updated_at
            ws[f"F{i}"] = job.rev_note
            ws[f"U{i}"] = job.c_no

        job.meta_data['parent'].reverse()
        if len(job.meta_data['parent']) > 0:
            p0 = job.meta_data['parent'][0]
            print(f"parint {job.meta_data['parent'][0]}")
            #print(f"parint {job.parent}")
            if p0:
                temp = Job.objects.get(pk = p0['id'])
                ws["A70"] = temp.version_code
                ws["C70"] = temp.updated_at
                ws["F70"] = temp.rev_note
                ws["U70"] = temp.c_no
            try:
                p1 = job.meta_data['parent'][1]
                print(f"parint {job.meta_data['parent'][0]}")
                #print(f"parint {job.parent}")
                if p1:
                    temp = Job.objects.get(pk = p1['id'])
                    ws["A71"] = temp.version_code
                    ws["C71"] = temp.updated_at
                    ws["F71"] = temp.rev_note
                    ws["U71"] = temp.c_no
            except Exception as e:
                pass

        furnance = jobfurnance.furnance
        print(f"values = {jobfurnance.furnance}")
        if furnance:
            print(f"furnance  = {furnance} {furnance.furtype} {furnance.furno}")
        doc_code = part.doc_code
        if doc_code is None:
            doc_code = "000"

        qrData = f"{part.doc_code}|{part.partname}|{part.designno}|{doc_code[0:3].zfill(4)}|{customer.name}|{furnance.furtype}|{furnance.furno}|{job.updated_at.strftime('%d-%B-%y')}|{str(job.version_code).zfill(2)}"
        print(f"qr = {qrData}")

        img = qrcode.make(qrData)
        thimg = img.resize((100, 100))
        ts = time.time()
        filename =  f"{int(ts)}.png"
        print(filename)
        fullPath = BASE_DIR + "/media/" + filename
        thimg.save(fullPath, format='PNG')
        qrimg = Image(fullPath)
        ws.add_image(qrimg, "Y5")
        '''
        if part.shtester1 is not None:
            try:
                const = ConstantMF.objects.get(data_code = part.shtester1. class_code=7)
                ws["B14"] = const.data1
            except:
                pass
        '''

        if part.shtester2 is not None:
            try:
                const = ConstantMF.objects.get(data_code = part.shtester2, class_code=7)
                ws["B14"] = const.data1
                ws["E14"] = part.low2
                ws["I14"] = part.up2
            except:
                pass

        if part.shtester3 is not None:
            try:
                const = ConstantMF.objects.get(data_code = part.shtester3, class_code=7)
                ws["B15"] = const.data1
                ws["E15"] = part.low3
                ws["I15"] = part.up3
            except ConstantMF.DoesNotExist:
                pass

        if part.chtester1 is not None:
            try:
                const = ConstantMF.objects.get(data_code = part.chtester1, class_code = 7)
                ws["N13"] = const.data1
                ws["Q13"] = part.chlow1
                ws["U13"] = part.chup1
            except ConstantMF.DoesNotExist:
                pass

        if part.chtester2 is not None:
            try:
                const = ConstantMF.objects.get(data_code = part.chtester2, class_code = 7)
                ws["N14"] = const.data1
                ws["Q14"] = part.chlow2
                ws["U14"] = part.chup2
            except ConstantMF.DoesNotExist:
                pass

        if part.othertester is not None:
            try:
                const = ConstantMF.objects.get(data_code = part.othertester, class_code = 7)
                ws["Q15"] = "{0} {1}-{2}".format(const.data1, part.olow1, part.oup1)
            except ConstantMF.DoesNotExist:
                pass

        if part.effpoint1 is not None and part.effpoint1 != 0:
            try:
                effpoint  = ConstantMF.objects.get(data_code = part.effpoint1, class_code = 8)
                ws["E17"] = effpoint.data
                ws["I17"] = part.effcriterion1
                ws["M17"] = part.eff_low1
                ws["Q17"] = part.eff_up1
            except ConstantMF.DoesNotExist:
                pass

        if part.effpoint2 is not None and part.effpoint2 != 0:
            try:
                effpoint2  = ConstantMF.objects.get(data_code = part.effpoint2, class_code = 8)
                ws["E18"] = effpoint2.data
                ws["I18"] = part.effcriterion2
                ws["M18"] = part.eff_low2
                ws["Q18"] = part.eff_up2
            except ConstantMF.DoesNotExist:
                pass

        ws["E19"] = part.total_measures
        ws["I19"] = part.total_low


        values = jobfurnance.values
        if values is not None:
            if 'safety' in values:
                ws['E20'] = 'R' if 'helmet' in values['safety'] else ''
                ws['I20'] = 'R' if 'glass' in values['safety'] else ''
                ws['M20'] = 'R' if 'ear_fill' in values['safety'] else ''
                ws['Q20'] = 'R' if 'nose' in values['safety'] else ''
                ws['U20'] = 'R' if 'eam' in values['safety'] else ''
                ws['X20'] = 'R' if 'arm' in values['safety'] else ''
                ws['AB20'] = 'R' if 'back' in values['safety'] else ''

                ws['E21'] = 'R' if 'safety_shoes' in values['safety'] else ''
                ws['I21'] = 'R' if 'ear_cover' in values['safety'] else ''
                ws['M21'] = 'R' if 'belt' in values['safety'] else ''
                ws['Q21'] = 'R' if 'glove' in values['safety'] else ''
                ws['U21'] = 'R' if 'glove_c' in values['safety'] else ''
                ws['X21'] = 'R' if 'other' in values['safety'] else ''

            if jobfurnance.template_name == "conveyor":
                print("create conveyor")
                conveyorTemplate(ws,jobfurnance)
            if jobfurnance.template_name == "batch-tvq":
                #tvqTemplate(ws, jobfurnance)
                genTemplate(ws, jobfurnance, tvqBody, img)

            if jobfurnance.template_name == "batch-tvt":
                genTemplate(ws, jobfurnance, tvtBody)

            if jobfurnance.template_name == "batch-tfq_tft":
                genTemplate(ws, jobfurnance, tfqBody)


            if jobfurnance.template_name == "batch-tgc_trgc":
                genTemplate(ws, jobfurnance, tgcBody)

            if jobfurnance.template_name == "batch-tbt":
                genTemplate(ws, jobfurnance, tbtBody)

            if jobfurnance.template_name == "batch-tih":
                genTemplate(ws, jobfurnance, tihBody)

            if jobfurnance.template_name == "batch-tsz":
                genTemplate(ws, jobfurnance, tszBody, False)

            if jobfurnance.template_name == "batch-tsp_tsap":
                genTemplate(ws, jobfurnance, tspBody, False)

def prevWash(ws, jobfurnance):
    values = jobfurnance.values
    ws['B25'] = 'P' if values.get('prev_wash_notneed', False) else ''
    ws['E25'] = 'P' if values.get('prev_wash_tow', False) else ''
    ws['K25'] = 'P' if values.get("prev_wash_tbo", False) else ''
    ws['q25'] = 'P' if values.get("prev_wash_tbt", False) else ''

    #i24 boutan
    ws["x24"] = 'P' if values.get("boutan", False)  else ''
    #n24 handsetting
    ws["ac24"] = 'P' if values.get("hand_setting", False) else ''
    #i28 rust need
    ws["x29"] = 'P' if values.get("rust_prevention", False) else ''
    #l29 rust comment
    ws['aa29'] = values.get("rust_prevention_comment", "")

    ws["e25"] = 'P' if values.get('tow', False) else ''
    ws["h26"] = values.get("tow_ptn", "")
    ws["h27"] = values.get("tow_temp", "")
    ws["h28"] = values.get("tow_time", "")

    ws["k25"] = 'P' if values.get('tbo', False) else ''
    ws["n26"] = values.get("tbo_ptn", "")
    ws["n27"] = values.get("tbo_temp", "")
    ws["n28"] = values.get("tbo_time", "")

    ws["q25"] = 'P' if values.get('tbt', False) else ''
    ws["t26"] = values.get("tbt_ptn", "")
    ws["t27"] = values.get("tbt_temp", "")
    ws["t28"] = values.get("tbt_time", "")

def inspectionPart(ws, jobfurnance):
    values = jobfurnance.values

    inspection = values.get("inspection", {})

    ws["a58"] = jobfurnance.furnance.checkq_1
    ws["q58"] = jobfurnance.furnance.checkt_2

    ws["d60"] = inspection.get("q_s_pd", "")
    #f60 q_s_std
    ws["f60"] = inspection.get("q_s_std", "")
    #h60 q_std_low
    ws["h60"] = inspection.get("q_s_std_low", "")
    #k60 q_std_hi
    ws["k60"] = inspection.get("q_s_std_hi", "")
    #n60 q_s_sent_qa
    ws["n60"] = inspection.get("q_s_sent_qa", "")

    #..61 and use c

    ws["d61"] = inspection.get("q_c_pd", "")
    #f60 q_s_std
    ws["f61"] = inspection.get("q_c_std", "")
    #h60 q_std_low
    ws["h61"] = inspection.get("q_c_std_low", "")
    #k60 q_std_hi
    ws["k61"] = inspection.get("q_c_std_hi", "")

    #t60, v, x, aa, ad
    ws["t60"] = inspection.get("t_s_pd", "")
    ws["v60"] = inspection.get("t_s_std", "")
    ws["x60"] = inspection.get("t_s_std_low", "")
    ws["aa60"] = inspection.get("t_s_std_hi", "")
    ws["ad60"] = inspection.get("t_s_sent_qa", "")

    ws["t61"] = inspection.get("t_c_pd", "")
    ws["v61"] = inspection.get("t_c_std", "")
    ws["x61"] = inspection.get("t_c_std_low", "")
    ws["aa61"] = inspection.get("t_c_std_hi", "")


    jobPhotos = jobfurnance.jobfurnancefile_set.filter(group = "Tempering")
    jobQPhotos = jobfurnance.jobfurnancefile_set.filter(group = "Quenching")

    if len(jobPhotos) > 0:
        try:
            img = Image(BASE_DIR + jobPhotos[0].file.url)
            ws.add_image(img, "A62")
        except:
            pass

        try:
            img = Image(BASE_DIR + jobPhotos[1].file.url)
            ws.add_image(img, "F62")
        except:
            pass

        try:
            img = Image(BASE_DIR + jobPhotos[2].file.url)
            ws.add_image(img, "K62")
        except:
            pass

    if len(jobQPhotos) > 0:

        try:
            img2 = Image(BASE_DIR + jobQPhotos[0].file.url)
            ws.add_image(img2, "Q62")
        except:
            pass

        try:
            img2 = Image(BASE_DIR + jobQPhotos[1].file.url)
            ws.add_image(img2, "V62")
        except:
            pass


        try:
            img2 = Image(BASE_DIR + jobQPhotos[2].file.url)
            ws.add_image(img2, "AA62")
        except:
            pass
def tvqBody(ws, jobfurnance):
    values = jobfurnance.values
    jf = jobfurnance
    ws["b30"] = jf.furnance.furnance
    ws["d31"] = values.get('ptn', '')
    for k,v in values.items():
        if k == "temp":
            ws["f32"] = get_value(v, "qcol1")
            ws["i32"] = get_value(v, "qcol2")
            ws["l32"] = get_value(v, "qcol3")
            ws["o32"] = get_value(v, "qcol4")
            ws["r32"] = get_value(v, "qcol5")
            ws["u32"] = get_value(v, "qcol6")
            ws["x32"] = get_value(v, "qcol7")
            ws["aa32"] = get_value(v, "qcol8")

            ws["f38"] = get_value(v, "tcol1")
            ws["i38"] = get_value(v, "tcol2")
            ws["l38"] = get_value(v, "tcol3")
            ws["o38"] = get_value(v, "tcol4")
            ws["r38"] = get_value(v, "tcol5")
            ws["u38"] = get_value(v, "tcol6")

        if k == "time":
            ws["f33"] = get_value(v, "qcol1")
            ws["i33"] = get_value(v, "qcol2")
            ws["l33"] = get_value(v, "qcol3")
            ws["o33"] = get_value(v, "qcol4")
            ws["r33"] = get_value(v, "qcol5")
            ws["u33"] = get_value(v, "qcol6")
            ws["x33"] = get_value(v, "qcol7")
            ws["aa33"] = get_value(v, "qcol8")

            ws["f39"] = get_value(v, "tcol1")
            ws["i39"] = get_value(v, "tcol2")
            ws["l39"] = get_value(v, "tcol3")
            ws["o39"] = get_value(v, "tcol4")
            ws["r39"] = get_value(v, "tcol5")
            ws["u39"] = get_value(v, "tcol6")

        if k == "pressure":
            ws["f34"] = get_value(v, "qcol1")
            ws["i34"] = get_value(v, "qcol2")
            ws["l34"] = get_value(v, "qcol3")
            ws["o34"] = get_value(v, "qcol4")
            ws["r34"] = get_value(v, "qcol5")
            ws["u34"] = get_value(v, "qcol6")
            ws["x34"] = get_value(v, "qcol7")
            ws["aa34"] = get_value(v, "qcol8")

            ws["f40"] = get_value(v, "tcol1")
            ws["i40"] = get_value(v, "tcol2")
            ws["l40"] = get_value(v, "tcol3")
            ws["o40"] = get_value(v, "tcol4")
            ws["r40"] = get_value(v, "tcol5")
            ws["u40"] = get_value(v, "tcol6")

        if k == "h2flow":
            ws["f41"] = get_value(v, "tcol1")
            ws["i41"] = get_value(v, "tcol2")
            ws["l41"] = get_value(v, "tcol3")
            ws["o41"] = get_value(v, "tcol4")
            ws["r41"] = get_value(v, "tcol5")
            ws["u41"] = get_value(v, "tcol6")

def tfqBody(ws, jobfurnance):
    values = jobfurnance.values
    jf = jobfurnance
    ws["b30"] = ws["b36"] = jf.furnance.furnance
    ws["d31"] = values.get('qptn', '')
    ws["d37"] = values.get('tptn', '')

    ws["f31"] = jf.furnance.qcol1
    ws["i31"] = jf.furnance.qcol2
    ws["l31"] = jf.furnance.qcol3
    ws["o31"] = jf.furnance.qcol4
    ws["r31"] = jf.furnance.qcol5
    ws["u31"] = jf.furnance.qcol6
    ws["x31"] = jf.furnance.qcol7
    ws["aa31"] = jf.furnance.qcol8

    ws["f37"] = jf.furnance.tcol1
    ws["i37"] = jf.furnance.tcol2
    ws["l37"] = jf.furnance.tcol3
    ws["o37"] = jf.furnance.tcol4
    ws["r37"] = jf.furnance.tcol5
    ws["u37"] = jf.furnance.tcol6

    for k,v in values.items():

        if k == "temp":
            ws["f32"] = get_value(v, "qcol1")
            ws["i32"] = get_value(v, "qcol2")
            ws["l32"] = get_value(v, "qcol3")
            ws["o32"] = get_value(v, "qcol4")
            ws["r32"] = get_value(v, "qcol5")
            ws["u32"] = get_value(v, "qcol6")
            ws["x32"] = get_value(v, "qcol7")
            ws["aa32"] = get_value(v, "qcol8")

            ws["f38"] = get_value(v, "tcol1")
            ws["i38"] = get_value(v, "tcol2")
            ws["l38"] = get_value(v, "tcol3")
            ws["o38"] = get_value(v, "tcol4")
            ws["r38"] = get_value(v, "tcol5")
            ws["u38"] = get_value(v, "tcol6")

        if k == "time":
            ws["f33"] = get_value(v, "qcol1")
            ws["i33"] = get_value(v, "qcol2")
            ws["l33"] = get_value(v, "qcol3")
            ws["o33"] = get_value(v, "qcol4")
            ws["r33"] = get_value(v, "qcol5")
            ws["u33"] = get_value(v, "qcol6")
            ws["x33"] = get_value(v, "qcol7")
            ws["aa33"] = get_value(v, "qcol8")

            ws["f39"] = get_value(v, "tcol1")
            ws["i39"] = get_value(v, "tcol2")
            ws["l39"] = get_value(v, "tcol3")
            ws["o39"] = get_value(v, "tcol4")
            ws["r39"] = get_value(v, "tcol5")
            ws["u39"] = get_value(v, "tcol6")

        if k == "pressure":
            ws["f34"] = get_value(v, "qcol1")
            ws["i34"] = get_value(v, "qcol2")
            ws["l34"] = get_value(v, "qcol3")
            ws["o34"] = get_value(v, "qcol4")
            ws["r34"] = get_value(v, "qcol5")
            ws["u34"] = get_value(v, "qcol6")
            ws["x34"] = get_value(v, "qcol7")
            ws["aa34"] = get_value(v, "qcol8")

            ws["f40"] = get_value(v, "tcol1")
            ws["i40"] = get_value(v, "tcol2")
            ws["l40"] = get_value(v, "tcol3")
            ws["o40"] = get_value(v, "tcol4")
            ws["r40"] = get_value(v, "tcol5")
            ws["u40"] = get_value(v, "tcol6")

        if k == "h2flow":
            ws["f41"] = get_value(v, "tcol1")
            ws["i41"] = get_value(v, "tcol2")
            ws["l41"] = get_value(v, "tcol3")
            ws["o41"] = get_value(v, "tcol4")
            ws["r41"] = get_value(v, "tcol5")
            ws["u41"] = get_value(v, "tcol6")

def tgcBody(ws, jobfurnance):
    values = jobfurnance.values
    jf = jobfurnance
    ws["b30"] = jf.furnance.furnance
    ws["d31"] = values.get('qptn', '')
    rowFill(ws, ["f", "i", "l", "o", "r", "u"], 31, jf.furnance, ["qcol1", "qcol2", "qcol3", "qcol4", "qcol5", "qcol6"])
    for k,v in values.items():

        if k == "temp":
            ws["f32"] = get_value(v, "qcol1")
            ws["i32"] = get_value(v, "qcol2")
            ws["l32"] = get_value(v, "qcol3")
            ws["o32"] = get_value(v, "qcol4")
            ws["r32"] = get_value(v, "qcol5")
            ws["u32"] = get_value(v, "qcol6")
            ws["x32"] = get_value(v, "qcol7")
            ws["aa32"] = get_value(v, "qcol8")


        if k == "cp":
            ws["f33"] = get_value(v, "qcol1")
            ws["i33"] = get_value(v, "qcol2")
            ws["l33"] = get_value(v, "qcol3")
            ws["o33"] = get_value(v, "qcol4")
            ws["r33"] = get_value(v, "qcol5")
            ws["u33"] = get_value(v, "qcol6")
            ws["x33"] = get_value(v, "qcol7")
            ws["aa33"] = get_value(v, "qcol8")


        if k == "co2":
            ws["f34"] = get_value(v, "qcol1")
            ws["i34"] = get_value(v, "qcol2")
            ws["l34"] = get_value(v, "qcol3")
            ws["o34"] = get_value(v, "qcol4")
            ws["r34"] = get_value(v, "qcol5")
            ws["u34"] = get_value(v, "qcol6")
            ws["x34"] = get_value(v, "qcol7")
            ws["aa34"] = get_value(v, "qcol8")

        #cols = ["f", "i", "l", "o", "r", "u", "x", "aa"]
        cols = ["f", "i", "l", "o", "r", "u"]
        #qcols = ["qcol1", "qcol2", "qcol3", "qcol4", "qcol5", "qcol6", "qcol6", "qcol7", "qcol8"]
        qcols = ["qcol1", "qcol2", "qcol3", "qcol4", "qcol5", "qcol6", "qcol6"]
        if k == "nh3":
            rowFill(ws, cols , 35, v, qcols)
        if k == "time":
            rowFill(ws, cols , 36, v, qcols)
        if k == "pressure":
            rowFill(ws, cols , 37, v, qcols)
        if k == "rxgas":
            rowFill(ws, cols , 38, v, qcols)
        if k == "speed":
            rowFill(ws, ["f"], 39, v, ["qcol1"])

        if k == "oiltemp":
            rowFill(ws, ["f", "o", "u"] , 40, v, ["qcol1", "qcol2", "qcol3"])

        if k == "agitator":
            rowFill(ws, ["i", "o"] , 41, v, ["qcol1", "qcol2"])


def tbtBody(ws, jobfurnance):
    values = jobfurnance.values
    jf = jobfurnance
    ws["b30"] = jf.furnance.furnance
    ws["d31"] = values.get('qptn', '')
    rowFill(ws, ["f", "i"], 31, jf.furnance, ["tcol1", "tcol2"])

    cols = ["f", "i"]
    vcols = ["tcol1", "tcol2"]

    for k,v in values.items():
        if k == "temp":
            rowFill(ws, cols , 32, v, vcols)

        if k == "time":
            rowFill(ws, cols , 33, v, vcols)

        if k == "speed":
            rowFill(ws, ["f"], 34, v, ["tcol1"])

def tszBody(ws, jobfurnance):
    values = jobfurnance.values
    jf = jobfurnance
    ws["b30"] = jf.furnance.furnance
    ws["d31"] = values.get('qptn', '')
    rowFill(ws, ["f"], 31, jf.furnance, ["qcol1"])
    rowFill(ws, ["i"], 31, jf.furnance, ["qcol2"])

    cols = ["f", "i"]
    vcols = ["qcol1", "qcol2"]

    rowFill(ws, ["f"], 32, values['temp'], ["qcol1"])
    rowFill(ws, ["i"], 32, values['temp'], ["qcol2"])
    rowFill(ws, ["f"], 33, values["time"], ["qcol1"])
    rowFill(ws, ["i"], 33, values["time"], ["qcol2"])

    '''
    for k,v in values.items():
        if k == "temp":
            rowFill(ws, cols , 32, v, vcols)

        if k == "time":
            rowFill(ws, cols , 33, v, vcols)
    '''


def tihBody(ws, jobfurnance):
    values = jobfurnance.values
    jf = jobfurnance
    ws["b30"] = jf.furnance.furnance
    ws["d31"] = values.get('qptn', '')

    rowFill(ws, ["b"], 32, jf.furnance, ["qcol1"])
    rowFill(ws, ["b"], 33, jf.furnance, ["qcol2"])
    rowFill(ws, ["b"], 34, jf.furnance, ["qcol3"])
    rowFill(ws, ["b"], 35, jf.furnance, ["qcol4"])
    rowFill(ws, ["b"], 36, jf.furnance, ["qcol5"])
    rowFill(ws, ["b"], 37, jf.furnance, ["qcol6"])
    rowFill(ws, ["b"], 38, jf.furnance, ["qcol7"])
    rowFill(ws, ["b"], 39, jf.furnance, ["qcol8"])

    rowFill(ws, ["b"], 40, jf.furnance, ["tcol1"])
    rowFill(ws, ["b"], 41, jf.furnance, ["tcol2"])
    rowFill(ws, ["b"], 42, jf.furnance, ["tcol3"])
    rowFill(ws, ["b"], 43, jf.furnance, ["tcol4"])
    rowFill(ws, ["b"], 44, jf.furnance, ["tcol5"])
    rowFill(ws, ["b"], 45, jf.furnance, ["tcol6"])

    rowFill(ws, ["f"], 32, values['induction'], ["qcol1"])
    rowFill(ws, ["f"], 33, values['induction'], ["qcol2"])
    rowFill(ws, ["f"], 34, values['induction'], ["qcol3"])
    rowFill(ws, ["f"], 35, values['induction'], ["qcol4"])
    rowFill(ws, ["f"], 36, values['induction'], ["qcol5"])
    rowFill(ws, ["f"], 37, values['induction'], ["qcol6"])
    rowFill(ws, ["f"], 38, values['induction'], ["qcol7"])
    rowFill(ws, ["f"], 39, values['induction'], ["qcol8"])
    rowFill(ws, ["f"], 40, values['induction'], ["tcol1"])
    rowFill(ws, ["f"], 41, values['induction'], ["tcol2"])
    rowFill(ws, ["f"], 42, values['induction'], ["tcol3"])
    rowFill(ws, ["f"], 43, values['induction'], ["tcol4"])
    rowFill(ws, ["f"], 44, values['induction'], ["tcol5"])
    rowFill(ws, ["f"], 45, values['induction'], ["tcol6"])

def tspBody(ws, jobfurnance):
    values = jobfurnance.values
    jf = jobfurnance
    ws["b30"] = jf.furnance.furnance
    ws["d31"] = values.get('qptn', '')

    rowFill(ws, ["b"], 32, jf.furnance, ["qcol1"])
    rowFill(ws, ["b"], 33, jf.furnance, ["qcol2"])
    rowFill(ws, ["b"], 34, jf.furnance, ["qcol3"])
    rowFill(ws, ["b"], 35, jf.furnance, ["qcol4"])
    rowFill(ws, ["b"], 36, jf.furnance, ["qcol5"])
    rowFill(ws, ["b"], 37, jf.furnance, ["qcol6"])


    if 'press_temp' in values:
        rowFill(ws, ["f"], 32, values['press_temp'], ["qcol1"])
        rowFill(ws, ["f"], 33, values['press_temp'], ["qcol2"])
        rowFill(ws, ["f"], 34, values['press_temp'], ["qcol3"])
        rowFill(ws, ["f"], 35, values['press_temp'], ["qcol4"])
        rowFill(ws, ["f"], 36, values['press_temp'], ["qcol5"])
        rowFill(ws, ["f"], 37, values['press_temp'], ["qcol6"])



def rowFill(ws, chars, rowno, va, vk):
    for c in range(len(chars)):
        ws[f"{chars[c]}{rowno}"] = get_value(va, vk[c])



def tvtBody(ws, jobfurnance):
    values = jobfurnance.values
    jf = jobfurnance
    ws["b30"] = jf.furnance.furnance
    ws["d31"] = values.get('ptn', '')

    ws["f31"] = jf.furnance.tcol1
    ws["i31"] = jf.furnance.tcol2
    ws["l31"] = jf.furnance.tcol3
    ws["o31"] = jf.furnance.tcol4
    ws["r31"] = jf.furnance.tcol5
    ws["u31"] = jf.furnance.tcol6

    for k,v in values.items():
        if k == "temp":
            ws["f32"] = get_value(v, "tcol1")
            ws["i32"] = get_value(v, "tcol2")
            ws["l32"] = get_value(v, "tcol3")
            ws["o32"] = get_value(v, "tcol4")
            ws["r32"] = get_value(v, "tcol5")
            ws["u32"] = get_value(v, "tcol6")

        if k == "time":
            ws["f33"] = get_value(v, "tcol1")
            ws["i33"] = get_value(v, "tcol2")
            ws["l33"] = get_value(v, "tcol3")
            ws["o33"] = get_value(v, "tcol4")
            ws["r33"] = get_value(v, "tcol5")
            ws["u33"] = get_value(v, "tcol6")

        if k == "pressure":
            ws["f34"] = get_value(v, "tcol1")
            ws["i34"] = get_value(v, "tcol2")
            ws["l34"] = get_value(v, "tcol3")
            ws["o34"] = get_value(v, "tcol4")
            ws["r34"] = get_value(v, "tcol5")
            ws["u34"] = get_value(v, "tcol6")

        if k == "h2flow":
            ws["f35"] = get_value(v, "tcol1")
            ws["i35"] = get_value(v, "tcol2")
            ws["l35"] = get_value(v, "tcol3")
            ws["o35"] = get_value(v, "tcol4")
            ws["r35"] = get_value(v, "tcol5")
            ws["u35"] = get_value(v, "tcol6")

def genTemplate(ws, jobfurnance, bodygen, isPrevWash = True):
    values = jobfurnance.values

    if isPrevWash == True:
        prevWash(ws, jobfurnance)

    bodygen(ws,jobfurnance)
    inspectionPart(ws, jobfurnance)


def conveyorTemplate(ws, jobfurnance):
    values = jobfurnance.values


    ws['B25'] = 'P' if values.get('prev_wash_notneed', False) else ''
    ws['E25'] = 'P' if values.get('prev_wash_trw', False) else ''
    ws['E26'] = 'P' if values.get("prev_wash_tow", False) else ''
    ws['E27'] = 'P' if values.get("prev_wash_tbo", False) else ''
    ws['E28'] = 'P' if values.get("prev_wash_tbt", False) else ''

    #i24 boutan
    ws["i24"] = 'P' if values.get("boutan", False)  else ''
    #n24 handsetting
    ws["n24"] = 'P' if values.get("hand_setting", False) else ''
    #i28 rust need
    ws["i28"] = 'P' if values.get("rust_prevention", False) else ''
    #l29 rust comment
    ws['l29'] = values.get("rust_prevention_comment", "")

    #s24 auto
    ws["s24"] = 'P' if values.get("hopper_auto", False) else ''
    #y25 chargerate
    ws["y25"] = values.get("chagerate_kg", "")
    #y26 charttime
    ws["y26"] = values.get("charttime", "")
    #y27 totalweight
    ws["y27"] = values.get("totalweight", "")
    #y28 virbate level
    ws["y28"] = values.get("vibration_level", "")
    #y29 dumming shutter
    ws["y29"] = values.get("dumming_shutter_high", "")


    jf = jobfurnance

    #b30 main furnance
    ws['b30'] = jf.furnance.furnance
    #f31, i31, l31, o31, r31, u31 furnance qcol1
    ws['f31'] = jf.furnance.qcol1
    ws['i31'] = jf.furnance.qcol2
    ws['l31'] = jf.furnance.qcol3
    ws['o31'] = jf.furnance.qcol4
    ws['r31'] = jf.furnance.qcol5
    ws['u31'] = jf.furnance.qcol6

    #b43 temp furnance
    temperingNo = jf.furnance.furnance.split("-")[1]
    ws["b43"] = "TRT-"+temperingNo
    #f44, i44, l44, o44, r44, u44 furnance tcol1
    ws["f44"] = jf.furnance.tcol1
    ws["i44"] = jf.furnance.tcol2
    ws["l44"] = jf.furnance.tcol3
    ws["o44"] = jf.furnance.tcol4
    ws["r44"] = jf.furnance.tcol5
    ws["u44"] = jf.furnance.tcol6


    #q temp
    vcols = ["f", "i", "l", "o", "r", "u"]

    #furnance params dict
    paramRows = {"temp": 32, "cp":33, "nh3": 34, "rxgas":35, "speed": 36, "oiltemp": 37, "agitator": 38, "feedpattern": 39, "coolingfan":40, "jetjump":41, "jetjump2":42}

    tRows = {"temp": 45, "time": 46, "speed":47, "agitator": 48}

    #tbt
    if 'tbt' in jf.values:
        if 'temp' in jf.values['tbt']:
            ws["ac44"] = jf.values['tbt']['temp']
        if 'time' in jf.values['tbt']:
            ws["ac45"] = jf.values['tbt']['time']


    for k,v in jf.values.items():
        if k == "cp":
            # both q , t
            qr = str(paramRows[k])
            #tr =  tRows[k]
            ws["f"+qr] = get_value(v, "qcol1")
            ws["i"+qr] = get_value(v, "qcol2")
            ws["l"+qr] = get_value(v, "qcol3")
            ws["o"+qr] = get_value(v , "qcol4")
            ws["r"+qr] = get_value(v , "qcol5")
            ws["u"+qr] = get_value(v, "qcol6")

        if k == "temp":
            qr = str(paramRows[k])
            tr =  str(tRows[k])
            ws["f"+qr] = get_value(v, "qcol1")
            ws["i"+qr] = get_value(v, "qcol2")
            ws["l"+qr] = get_value(v, "qcol3")
            ws["o"+qr] = get_value(v, "qcol4")
            ws["r"+qr] = get_value(v, "qcol5")
            ws["u"+qr] = get_value(v, "qcol6")


            ws["f"+tr] = get_value(v, "tcol1")
            ws["i"+tr] = get_value(v, "tcol2")
            ws["l"+tr] = get_value(v, "tcol3")
            ws["o"+tr] = get_value(v, "tcol4")
            ws["r"+tr] = get_value(v, "tcol5")
            ws["u"+tr] = get_value(v, "tcol6")

        if k == "nh3":
            qr = str(paramRows[k])
            #tr =  tRows[k]
            ws["f"+qr] = get_value(v, "qcol1")
            ws["i"+qr] = get_value(v, "qcol2")
            ws["l"+qr] = get_value(v, "qcol3")
            ws["o"+qr] = get_value(v , "qcol4")
            ws["r"+qr] = get_value(v , "qcol5")
            ws["u"+qr] = get_value(v, "qcol6")

        if k == "rxgas":
            qr = str(paramRows[k])
            #tr =  tRows[k]
            ws["f"+qr] = get_value(v, "qcol")

        if k == "speed":
            qr = str(paramRows[k])
            tr =  str(tRows[k])
            ws["f"+qr] = get_value(v, "qcol")
            ws["f"+tr] = get_value(v, "tcol")

        if k == "oiltemp":
            qr = str(paramRows[k])
            #tr =  tRows[k]
            ws["o"+qr] = get_value(v, "qcol")

        if k == "agitator":
            qr = str(paramRows[k])
            tr =  str(tRows[k])
            ws["f"+qr] = get_value(v, "qcol")
            ws["f"+tr] = get_value(v, "tcol1")
            ws["q"+tr] = get_value(v, "tcol2")

        if k == "feedpattern":
            qr = str(paramRows[k])
            #tr =  tRows[k]
            ws["f"+qr] = get_value(v, "qcol")

        if k == "coolingfan":
            qr = str(paramRows[k])
            #tr =  tRows[k]
            ws["i"+qr] = get_value(v, "qcol1")
            ws["r"+qr] = get_value(v, "qcol2")

        if k == "jetjump":
            qr = str(paramRows[k])
            #tr =  tRows[k]
            ws["f"+qr] = get_value(v, "qcol1")
            ws["o"+qr] = get_value(v, "qcol2")

        if k == "time":
            ws["f46"] = get_value(v, "tcol1")
            ws["i46"] = get_value(v, "tcol2")
            ws["l46"] = get_value(v, "tcol3")
            ws["o46"] = get_value(v, "tcol4")
            ws["r46"] = get_value(v, "tcol5")
            ws["u46"] = get_value(v, "tcol6")

        if k == "jump_on":
            if v == True:
                ws["f41"] = "R"
            else:
                ws["o41"] = "R"

    #inspection part
    #job object
    #d60 q_s_pd
    inspection = values.get("inspection", {})

    ws["a58"] = jf.furnance.checkq_1
    ws["q58"] = jf.furnance.checkt_2

    ws["d60"] = inspection.get("q_s_pd", "")
    #f60 q_s_std
    ws["f60"] = inspection.get("q_s_std", "")
    #h60 q_std_low
    ws["h60"] = inspection.get("q_s_std_low", "")
    #k60 q_std_hi
    ws["k60"] = inspection.get("q_s_std_hi", "")
    #n60 q_s_sent_qa
    ws["n60"] = inspection.get("q_s_sent_qa", "")

    #..61 and use c

    ws["d61"] = inspection.get("q_c_pd", "")
    #f60 q_s_std
    ws["f61"] = inspection.get("q_c_std", "")
    #h60 q_std_low
    ws["h61"] = inspection.get("q_c_std_low", "")
    #k60 q_std_hi
    ws["k61"] = inspection.get("q_c_std_hi", "")

    #t60, v, x, aa, ad
    ws["t60"] = inspection.get("t_s_pd", "")
    ws["v60"] = inspection.get("t_s_std", "")
    ws["x60"] = inspection.get("t_s_std_low", "")
    ws["aa60"] = inspection.get("t_s_std_hi", "")
    ws["ad60"] = inspection.get("t_s_sent_qa", "")

    ws["t61"] = inspection.get("t_c_pd", "")
    ws["v61"] = inspection.get("t_c_std", "")
    ws["x61"] = inspection.get("t_c_std_low", "")
    ws["aa61"] = inspection.get("t_c_std_hi", "")


    jobPhotos = jobfurnance.jobfurnancefile_set.filter(group = "Tempering")
    jobQPhotos = jobfurnance.jobfurnancefile_set.filter(group = "Quenching")

    if len(jobPhotos) > 0:
        try:
            img = Image(f"{BASE_DIR}{jobPhotos[0].file.url}")
            ws.add_image(img, "A62")
        except:
            pass

        try:
            img = Image(f"{BASE_DIR}{jobPhotos[1].file.url}")
            ws.add_image(img, "F62")
        except:
            pass

        try:
            img = Image(f"{BASE_DIR}{jobPhotos[2].file.url}")
            ws.add_image(img, "K62")
        except:
            pass

    if len(jobQPhotos) > 0:

        try:
            img2 = Image(f"{BASE_DIR}{jobQPhotos[0].file.url}")
            ws.add_image(img2, "Q62")
        except:
            pass

        try:
            img2 = Image(f"{BASE_DIR}{jobQPhotos[1].file.url}")
            ws.add_image(img2, "V62")
        except:
            pass


        try:
            img2 = Image(f"{BASE_DIR}{jobQPhotos[2].file.url}")
            ws.add_image(img2, "AA62")
        except:
            pass




def get_value(o, k):
    if isinstance(o, dict):
        if k in o:
            return o[k]
        else:
            return ""
    else:
        if hasattr(o, k):
            return getattr(o, k)
        else:
            return ""




from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic.list import ListView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.messages.views import SuccessMessageMixin
from .models import Furnance, Customer, PartMF, PartPhoto, ParamTemplate
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin
from django import forms
from .forms import PartPhotoForm, PartPhotoFormSet
from django.db import transaction
import django_filters
from django import forms
from django.core.files.storage import FileSystemStorage
from django.contrib import messages

# Create your views here.

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from django.db.models import Q

class UploadFileForm(forms.Form):
    #title = forms.CharField(max_length=50)
    importFile = forms.FileField(label="Part Import File")

class PartFilter(django_filters.FilterSet):
    partcode = django_filters.CharFilter(lookup_expr='icontains')
    doc_code = django_filters.CharFilter(lookup_expr='icontains')

    class Meta:
        model = PartMF
        fields = ['partcode', 'doc_code']

def index(request):
    return render(request, 'master/index.html')

class FurnanceFilter(django_filters.FilterSet):
    furnance = django_filters.CharFilter(lookup_expr='icontains')
    furtype = django_filters.CharFilter(lookup_expr='icontains')
    furno = django_filters.CharFilter(lookup_expr='icontains')

    class Meta:
        model = Furnance
        fields = ['furnance', 'furtype', 'furno']

class ParamTemplateListView(PermissionRequiredMixin, ListView):
    model = ParamTemplate
    paginate_by  = 50
    permission_required = 'master.view_paramtemplate'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['header'] = "ParamTemplate List"
        context['total'] = self.model.objects.count()
        #listFilter = FurnanceFilter(self.request.GET)
        #context['filter'] = listFilter
        return context

    def get_queryset(self):
        object_list =  self.model.objects.all().order_by('-id')
        #listFilter = FurnanceFilter(self.request.GET, queryset=object_list)
        return object_list

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('page_size', self.paginate_by)

class ParamTemplateCreateView(PermissionRequiredMixin, SuccessMessageMixin,CreateView):
    model = ParamTemplate
    fields = "__all__"
    permission_required = 'master.add_paramtemplate'
    success_message = "%(paramtemplate)s was created successfully"

    def get_form(self):
        form = super(ParamTemplateCreateView, self).get_form()
        return form

class ParamTemplateEditView(PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = ParamTemplate
    fields = "__all__"
    permission_required = 'master.change_paramtemplate'
    success_message = "%(paramtemplate)s was updated successfully"

    '''
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['furnance_list'] = ParamTemplate.objects.all()
        return context
    '''

    def get_form(self):
        form = super(ParamTemplateEditView, self).get_form()
        return form

    def post(self,request, **kwargs):
        if request.POST.get('clone'):
            print("clone !!")

        return super(ParamTemplateEditView, self).post(request, **kwargs)

class ParamTemplateDeleteView(PermissionRequiredMixin, DeleteView):
    model = ParamTemplate
    success_url = reverse_lazy("paramtemplate-list")

    permission_required = 'master.delete_paramtemplate'

    def get(self, *args, **kwargs):
        return self.post(*args, **kwargs)

class FurnanceListView(PermissionRequiredMixin, ListView):
    model = Furnance
    paginate_by  = 50
    permission_required = 'master.view_furnance'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['header'] = "Furnance List"
        context['total'] = self.model.objects.count()
        listFilter = FurnanceFilter(self.request.GET)
        context['filter'] = listFilter
        return context

    def get_queryset(self):
        object_list =  self.model.objects.all().order_by('-id')
        listFilter = FurnanceFilter(self.request.GET, queryset=object_list)
        return listFilter.qs

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('page_size', self.paginate_by)

class FurnanceCreateView(PermissionRequiredMixin, SuccessMessageMixin,CreateView):
    model = Furnance
    fields = "__all__"
    permission_required = 'master.add_furnance'
    success_message = "%(furnance)s was created successfully"

    def get_form(self):
        form = super(FurnanceCreateView, self).get_form()
        form.fields['params'].widget = forms.HiddenInput()
        form.fields['tparams'].widget = forms.HiddenInput()
        return form

class FurnanceEditView(PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Furnance
    fields = "__all__"
    permission_required = 'master.change_furnance'
    success_message = "%(furnance)s was updated successfully"

    '''
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['furnance_list'] = Furnance.objects.all()
        return context
    '''

    def get_form(self):
        form = super(FurnanceEditView, self).get_form()
        form.fields['params'].widget = forms.HiddenInput()
        form.fields['tparams'].widget = forms.HiddenInput()
        return form

    def post(self,request, **kwargs):
        if request.POST.get('clone'):
            print("clone !!")

        return super(FurnanceEditView, self).post(request, **kwargs)

def clone_furnance(request, pk):
    src = Furnance.objects.get(pk = pk)
    src.furnance = src.furnance + " Clone"
    src.pk = None
    src.save()

    return redirect("furnance-edit", pk = src.pk)

class FurnanceDeleteView(PermissionRequiredMixin, DeleteView):
    model = Furnance
    success_url = reverse_lazy("furnance-list")

    permission_required = 'master.delete_furnance'

    def get(self, *args, **kwargs):
        return self.post(*args, **kwargs)

class CustomerListView(ListView):
    model = Customer
    paginate_by  = 50

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['header'] = "Customer List"
        context['total'] = self.model.objects.count()
        return context

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('page_size', self.paginate_by)

class CustomerCreateView(SuccessMessageMixin,CreateView):
    model = Customer
    fields = "__all__"
    success_message = "%(name)s was created successfully"

class CustomerEditView(SuccessMessageMixin, UpdateView):
    model = Customer
    fields = "__all__"
    success_message = "%(name)s was updated successfully"


class CustomerDeleteView(DeleteView):
    model = Customer
    success_url = reverse_lazy("customer-list")

    def get(self, *args, **kwargs):
        return self.post(*args, **kwargs)

def importFile(request):
    print("Import file")
    uploadForm = UploadFileForm()
    uploaded_file_url = None

    if request.method == "POST":
        uploadForm = UploadFileForm(request.POST, request.FILES)
        if uploadForm.is_valid():
            fs = FileSystemStorage()
            filename = fs.save(uploadForm.cleaned_data['importFile'].name, uploadForm.cleaned_data['importFile'])
            uploaded_file_url = fs.url(filename)
            print("upload_file_url "+ uploaded_file_url )
            print(fs.path(filename))
            errors = []
            importPart(fs.path(filename), errors)
            if len(errors) == 0:
                messages.success(request, "Import Part Success")
            else:
                for e in errors:
                    messages.error(request, e)

            return redirect('importFile')

    return render(request, 'master/importFile.html', {'uploaded_file_url': uploaded_file_url, 'uploadForm': uploadForm})

def cell_val(o, letter):
    index = column_index_from_string(letter)
    v = o[index-1].value
    return v

def importPart(fn, errors):
    wb0 = load_workbook(fn,  read_only=True, data_only=True)
    ws = wb0['partmf']
    #models.Boxer.objects.all().delete()
    aord = ord('A')

    for row in ws.iter_rows(min_row=2):
        try:
            custom_no = cell_val(row, 'A')
            partcode = cell_val(row, 'B')
            doc_code = cell_val(row, 'C')
            treatcd = cell_val(row, 'D')
            designno = cell_val(row, 'E')
            partname = cell_val(row, 'F')
            size = cell_val(row, 'G')
            mat = cell_val(row, 'H')
            shtester1 = cell_val(row, 'I')
            low1 = cell_val(row, 'J')
            up1 = cell_val(row, 'K')
            shtester2 = cell_val(row, 'L')
            low2 = cell_val(row, 'M')
            up2 = cell_val(row, 'N')
            shtester3 = cell_val(row, 'O')
            low3 = cell_val(row, 'P')
            up3 = cell_val(row, 'Q')
            chtester1 = cell_val(row, 'R')
            chlow1 = cell_val(row, 'S')
            chup1 = cell_val(row, 'T')
            chtester2 = cell_val(row, 'U')
            chlow2 = cell_val(row, 'V')
            chup2 = cell_val(row, 'W')
            othertester = cell_val(row, 'X')
            olow1 = cell_val(row, 'Y')
            oup1 = cell_val(row, 'Z')
            effpoint1 = cell_val(row, 'AA')
            effcriterion1 = cell_val(row, 'AB')
            eff_low1 = cell_val(row, 'AC')
            eff_up1= cell_val(row, 'AD')

            effpoint2 = cell_val(row, 'AE')
            effcriterion2 = cell_val(row, 'AF')
            eff_low2 = cell_val(row, 'AG')
            eff_up2= cell_val(row, 'AH')

            total_measures = cell_val(row, 'AI')
            if total_measures == None:
                total_measures = 0

            total_low  = cell_val(row, 'AJ')
            if total_low == None:
                total_low = 0

            total_up  = cell_val(row, 'AK')
            if total_up == None:
                total_up = 0

            kgperpc = cell_val(row, 'AY')
            if kgperpc == None:
                kgperpc = 0

            rev_date = cell_val(row, 'BK')
            reg_date = cell_val(row, 'BJ')
            del_flag = cell_val(row, 'BL')
            qo_no = cell_val(row, 'BZ')
            if qo_no == None:
                qo_no = '0'
            #print("qo_no")
            #print(qo_no)


            #print("furno")
            #print(furnance)
            #print(furno)
            #print(qcol1)

            if partcode is not None:
                PartMF.objects.update_or_create(
                        partcode = partcode,
                        defaults = {
                            'custom_no': custom_no,
                            'doc_code': doc_code,
                            'treatcd': treatcd,
                            'designno': designno,
                            'partname':partname,
                            'size':  size,
                            'mat':mat,
                            'shtester1':shtester1,
                            'low1': low1,
                            'up1': up1,
                            'shtester2':shtester2,
                            'low2': low2,
                            'up2': up2,
                            'shtester3':shtester3,
                            'low3':low3,
                            'up3':up3,
                            'chtester1':chtester1,
                            'chlow1':chlow1,
                            'chup1':chup1,
                            'chtester2':chtester2,
                            'chlow2':chlow2,
                            'chup2':chup2,
                            'othertester':othertester,
                            'olow1':olow1,
                            'oup1':oup1,
                            'effpoint1':effpoint1,
                            'effcriterion1':effcriterion1,
                            'eff_low1':  eff_low1,
                            'eff_up1': eff_up1,
                            'effpoint2':  effpoint2,
                            'effcriterion2':  effcriterion2,
                            'eff_low2':  eff_low2,
                            'eff_up2': eff_up2,
                            'total_measures':  total_measures,
                            'total_low':  total_low,
                            'total_up':  total_up,
                            'rev_date':  rev_date,
                            'reg_date':  reg_date,
                            'del_flag':  del_flag,
                            'qo_no':  qo_no,
                            'kgperpc':  kgperpc
                            })
            print(".", end="")


        except Exception as e:
            print(str(e) + " " + str(row[0]))
            errors.append(str(e) + " " +str(row[0]))
    return True

class PartListView(ListView):
    model = PartMF
    paginate_by  = 50

    '''
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = self.model.objects.count()
        return context
    '''

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = self.model.objects.count()


        listFilter  = PartFilter(self.request.GET)
        context['filter'] = listFilter

        if self.request.POST:
            print("POST")

        return context

    def get_queryset(self):
        object_list =  self.model.objects.all().order_by('-id')
        listFilter = PartFilter(self.request.GET, queryset=object_list)
        return listFilter.qs

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('page_size', self.paginate_by)

class PartCreateView(SuccessMessageMixin,CreateView):
    model = PartMF
    fields = "__all__"
    success_message = "%(partcode)s was created successfully"


    def get_success_url(self):
        return reverse_lazy('part-edit',  kwargs = {'pk':self.object.id})

    def get_context_data(self, **kwargs):
        data = super(PartCreateView, self).get_context_data(**kwargs)
        if self.request.POST:
            data['partphotos'] = PartPhotoFormSet(self.request.POST, self.request.FILES)
        else:
            data['partphotos'] = PartPhotoFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        partphotos = context['partphotos']
        with transaction.atomic():
            self.object = form.save()
            if partphotos.is_valid():
                partphotos.instance = self.object
                partphotos.save()
        return super(PartCreateView, self).form_valid(form)

class PartEditView(SuccessMessageMixin, UpdateView):
    model = PartMF
    fields = "__all__"
    success_message = "%(partcode)s was updated successfully"


    def get_success_url(self):
        return reverse_lazy('part-edit',  kwargs = {'pk':self.object.id})

    def get_context_data(self, **kwargs):
        data = super(PartEditView, self).get_context_data(**kwargs)
        if self.request.POST:
            data['partphotos'] = PartPhotoFormSet(self.request.POST, self.request.FILES, instance=self.object)
        else:
            data['partphotos'] = PartPhotoFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        partphotos = context['partphotos']
        with transaction.atomic():
            self.object = form.save()
            if partphotos.is_valid():
                partphotos.instance = self.object
                partphotos.save()
        return super(PartEditView, self).form_valid(form)

class PartDeleteView(DeleteView):
    model = PartMF
    success_url = reverse_lazy("part-list")

    def get(self, *args, **kwargs):
        return self.post(*args, **kwargs)

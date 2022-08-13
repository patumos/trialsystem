from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from tohken import settings
import django
import os
from django.db.models import Q
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "tohken.settings")
django.setup()

from django.contrib.auth.models import User
#from django.core.management import setup_environ

from master import models
import datetime

import inspect
import sys
def whoami():
    frame = inspect.currentframe()
    return inspect.getframeinfo(frame).function

wb0 = load_workbook('masterdata.xlsx',  read_only=True, data_only=True)


def cell_val(o, letter):
    index = column_index_from_string(letter)
    v = o[index-1].value
    return v

def loadFurnance():
    models.Furnance.objects.all().delete()
    global wb0
    #print(wb2.sheetnames)
    print("Running : "+sys._getframe(  ).f_code.co_name)
    ws = wb0['furnance']
    #models.Boxer.objects.all().delete()
    aord = ord('A')

    for row in ws.iter_rows(min_row=3):
            furnance = cell_val(row, 'A')
            qcol1 = cell_val(row, 'B')
            qcol2 = cell_val(row, 'C')
            qcol3 = cell_val(row, 'D')
            qcol4 = cell_val(row, 'E')
            qcol5 = cell_val(row, 'F')
            qcol6 = cell_val(row, 'G')
            qcol7 = cell_val(row, 'H')
            qcol8 = cell_val(row, 'I')
            tcol1 = cell_val(row, 'J')
            tcol2 = cell_val(row, 'K')
            tcol3 = cell_val(row, 'L')
            tcol4 = cell_val(row, 'M')
            tcol5 = cell_val(row, 'N')
            tcol6 = cell_val(row, 'O')
            checkq_1 = cell_val(row, 'P')
            checkt_2 = cell_val(row, 'Q')
            coveyor_length = cell_val(row, 'R')
            max_weight_value = cell_val(row, 'S')
            max_weight_unit = cell_val(row, 'T')
            furtype = cell_val(row, 'U')
            furno = cell_val(row, 'V')

            #print("furno")
            #print(furnance)
            #print(furno)
            #print(qcol1)
            if furnance is not None:
                models.Furnance.objects.create(
                        furnance = furnance,
                        qcol1 = qcol1,
                        qcol2 = qcol2,
                        qcol3 = qcol3,
                        qcol4 =  qcol4,
                        qcol5 =  qcol5,
                        qcol6 =  qcol6,
                        qcol7 =  qcol7,
                        qcol8 =  qcol8,
                        tcol1 = tcol1,
                        tcol2 = tcol2,
                        tcol3 = tcol3,
                        tcol4 = tcol4,
                        tcol5 =  tcol5,
                        tcol6 =  tcol6,
                        checkq_1 =  checkq_1,
                        checkt_2 =  checkt_2,
                        coveyor_length =  coveyor_length,
                        max_weight_unit = max_weight_unit,
                        max_weight_value =  max_weight_value,
                        furtype =  furtype,
                        furno =  furno
                        )
            print(".", end="")

def loadCustomers():
    models.Customer.objects.all().delete()
    models.ConstantMF.objects.all().delete()
    global wb0
    #print(wb2.sheetnames)
    print("Running : "+sys._getframe(  ).f_code.co_name)
    ws = wb0['ConstantMF']
    #models.Boxer.objects.all().delete()
    aord = ord('A')

    for row in ws.iter_rows(min_row=2):
            code = cell_val(row, 'I')
            name = cell_val(row, 'J')

            #constant
            class_code = cell_val(row, 'A')
            data_code = cell_val(row, 'B')
            data1 = cell_val(row, 'C')
            data2 = cell_val(row, 'D')
            data3 = cell_val(row, 'E')
            total = cell_val(row, 'F')

            if class_code is not None:
                models.ConstantMF.objects.create(class_code = class_code, data_code = data_code, data1 = data1, data2 = data2, data3 = data3, total = total)
            if code is not  None:
                models.Customer.objects.create(
                        code  = code,
                        name =  name
                        )
            print(".", end="")



def loadParts():
    models.PartMF.objects.all().delete()
    global wb0
    #print(wb2.sheetnames)
    print("Running : "+sys._getframe(  ).f_code.co_name)
    ws = wb0['partmf']
    #models.Boxer.objects.all().delete()
    aord = ord('A')

    for row in ws.iter_rows(min_row=2):
        try:
            custom_no = cell_val(row, 'A')
            partcode = cell_val(row, 'B')
            doc_code = cell_val(row, 'C')
            treatcd = cell_val(row, 'D')
            designno = cell_val(row, 'E')
            partname = cell_val(row, 'F')
            size = cell_val(row, 'G')
            mat = cell_val(row, 'H')
            shtester1 = cell_val(row, 'I')
            low1 = cell_val(row, 'J')
            up1 = cell_val(row, 'K')
            shtester2 = cell_val(row, 'L')
            low2 = cell_val(row, 'M')
            up2 = cell_val(row, 'N')
            shtester3 = cell_val(row, 'O')
            low3 = cell_val(row, 'P')
            up3 = cell_val(row, 'Q')
            chtester1 = cell_val(row, 'R')
            chlow1 = cell_val(row, 'S')
            chup1 = cell_val(row, 'T')
            chtester2 = cell_val(row, 'U')
            chlow2 = cell_val(row, 'V')
            chup2 = cell_val(row, 'W')
            othertester = cell_val(row, 'X')
            olow1 = cell_val(row, 'Y')
            oup1 = cell_val(row, 'Z')
            effpoint1 = cell_val(row, 'AA')
            effcriterion1 = cell_val(row, 'AB')
            eff_low1 = cell_val(row, 'AC')
            eff_up1= cell_val(row, 'AD')

            effpoint2 = cell_val(row, 'AE')
            effcriterion2 = cell_val(row, 'AF')
            eff_low2 = cell_val(row, 'AG')
            eff_up2= cell_val(row, 'AH')

            total_measures = cell_val(row, 'AI')
            if total_measures == None:
                total_measures = 0

            total_low  = cell_val(row, 'AJ')
            if total_low == None:
                total_low = 0

            total_up  = cell_val(row, 'AK')
            if total_up == None:
                total_up = 0

            kgperpc = cell_val(row, 'AY')
            if kgperpc == None:
                kgperpc = 0

            rev_date = cell_val(row, 'BK')
            reg_date = cell_val(row, 'BJ')
            del_flag = cell_val(row, 'BL')
            qo_no = cell_val(row, 'BZ')
            if qo_no == None:
                qo_no = '0'
            #print("qo_no")
            #print(qo_no)


            #print("furno")
            #print(furnance)
            #print(furno)
            #print(qcol1)

            if partcode is not None:
                models.PartMF.objects.create(
                        partcode = partcode,
                        custom_no = custom_no,
                        doc_code = doc_code,
                        treatcd = treatcd,
                        designno =  designno,
                        partname = partname,
                        size =  size,
                        mat = mat,
                        shtester1 = shtester1,
                        low1  = low1,
                        up1 =  up1,
                        shtester2 = shtester2,
                        low2 =  low2,
                        up2 =  up2,
                        shtester3 = shtester3,
                        low3  =  low3,
                        up3 =  up3,
                        chtester1 =  chtester1,
                        chlow1  = chlow1,
                        chup1 =  chup1,
                        chtester2 =  chtester2,
                        chlow2 =  chlow2,
                        chup2 =  chup2,
                        othertester =  othertester,
                        olow1 =  olow1,
                        oup1 =  oup1,
                        effpoint1 =  effpoint1,
                        effcriterion1 =  effcriterion1,
                        eff_low1 =  eff_low1,
                        eff_up1 =  eff_up1,
                        effpoint2 =  effpoint2,
                        effcriterion2 =  effcriterion2,
                        eff_low2 =  eff_low2,
                        eff_up2 =  eff_up2,
                        total_measures =  total_measures,
                        total_low =  total_low,
                        total_up =  total_up,
                        rev_date =  rev_date,
                        reg_date =  reg_date,
                        del_flag =  del_flag,
                        qo_no =  qo_no,
                        kgperpc =  kgperpc
                        )
            print(".", end="")


        except Exception as e:
            print(str(e) + " " + str(row[0]))


def migrateData():

    #loadFurnance()
    loadParts()
    #loadCustomers()

def migratePartPhoto():
    parts = models.PartMF.objects.all()

    for p in parts:
        if p.doc_code is not None:
            try:
                pp = models.PartPhoto()
                pp.file = f"parts/{p.doc_code}.jpg"
                pp.part = p
                pp.save()
                print(".", end="")
            except:
                print("x", end="")



def migrateUser():
    global wb3
    ws = wb3['ส่วนกลาง']
    User.objects.filter(~Q(id=1)).delete()
    for row in ws.iter_rows(min_row=6, max_row=19):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = "กรุงเทพมหานคร"
            user.save()
        except Exception as e:
            print(str(e))

    ws = wb3['ผอ. ภาค 2']
    for row in ws.iter_rows(min_row=6, max_row=20):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))

    ws = wb3['ผอ. ภาค 1']
    for row in ws.iter_rows(min_row=6, max_row=19):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))

    ws = wb3['ผอ. ภาค 4']
    for row in ws.iter_rows(min_row=6, max_row=20):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))

    ws = wb3['ผอ. ภาค 5']
    for row in ws.iter_rows(min_row=6, max_row=21):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))

    ws = wb3['ผช.ภาค1']
    for row in ws.iter_rows(min_row=6, max_row=18):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))

    ws = wb3['ผช.ภาค2']
    for row in ws.iter_rows(min_row=6, max_row=19):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))

    ws = wb3['ผช.ภาค3']
    for row in ws.iter_rows(min_row=6, max_row=27):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))


    ws = wb3['ผช.ภาค4']
    for row in ws.iter_rows(min_row=6, max_row=19):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))

    ws = wb3['ผช.ภาค5']
    for row in ws.iter_rows(min_row=6, max_row=20):
        u = row[column_index_from_string('C')-1].value
        password = row[column_index_from_string('F')-1].value
        province = row[column_index_from_string('D')-1].value
        if u == "":
            continue
        if password == "":
            password = 1234

        role = row[column_index_from_string('G')-1].value.lower()
        fn = row[column_index_from_string('B')-1].value.split()
        try:
            user = User.objects.create_user(u, None, password)
            user.profile.staff_roles  = [role]
            user.is_active = True
            user.first_name = fn[0]
            user.last_name = fn[1]
            user.profile.province = province
            user.save()
        except Exception as e:
            print(str(e))


if __name__ == '__main__':
    #migrateData()
    action = sys.argv[1]
    if action == "user":
        migrateUser()
    if action == "data":
        migrateData()
    if action == "partphoto":
        migratePartPhoto()



